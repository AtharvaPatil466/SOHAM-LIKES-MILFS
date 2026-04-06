"""GST Returns export generators (GSTR-1, GSTR-3B format).

Generates Excel files in the format required for Indian GST filing.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


GST_SLABS = [0, 5, 12, 18, 28]
CATEGORY_RATES = {
    "Grocery": 5, "Dairy": 5, "Frozen": 12, "Snacks": 12,
    "Beverages": 12, "Personal Care": 18, "Cleaning": 18,
    "Baby Care": 12, "Bakery": 5, "Protein & Health": 18,
}


def generate_gstr1_excel(
    invoices: list[dict],
    date_from: str,
    date_to: str,
    gstin: str = "",
    store_name: str = "RetailOS Store",
) -> io.BytesIO:
    """Generate GSTR-1 format Excel (outward supplies summary).

    Sheets: B2B (business), B2C (consumer), HSN Summary.
    """
    wb = Workbook()

    # ── B2B Sheet (invoices to registered businesses) ──
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    _header_row(ws_b2b, 1, f"GSTR-1 B2B Invoices: {date_from} to {date_to}")
    _header_row(ws_b2b, 2, f"GSTIN: {gstin} | {store_name}")

    b2b_headers = ["GSTIN of Buyer", "Invoice No", "Invoice Date", "Invoice Value",
                   "Place of Supply", "Rate", "Taxable Value", "IGST", "CGST", "SGST"]
    _styled_headers(ws_b2b, 4, b2b_headers)

    row = 5
    b2b_invoices = [inv for inv in invoices if inv.get("buyer_gstin")]
    for inv in b2b_invoices:
        rate = inv.get("gst_rate", 18)
        taxable = inv.get("taxable_value", inv.get("total_amount", 0) / (1 + rate / 100))
        is_igst = inv.get("is_interstate", False)
        gst_amount = taxable * rate / 100

        ws_b2b.cell(row=row, column=1, value=inv.get("buyer_gstin", ""))
        ws_b2b.cell(row=row, column=2, value=inv.get("invoice_number", ""))
        ws_b2b.cell(row=row, column=3, value=inv.get("invoice_date", ""))
        ws_b2b.cell(row=row, column=4, value=round(inv.get("total_amount", 0), 2))
        ws_b2b.cell(row=row, column=5, value=inv.get("place_of_supply", ""))
        ws_b2b.cell(row=row, column=6, value=f"{rate}%")
        ws_b2b.cell(row=row, column=7, value=round(taxable, 2))
        ws_b2b.cell(row=row, column=8, value=round(gst_amount, 2) if is_igst else 0)
        ws_b2b.cell(row=row, column=9, value=0 if is_igst else round(gst_amount / 2, 2))
        ws_b2b.cell(row=row, column=10, value=0 if is_igst else round(gst_amount / 2, 2))
        row += 1

    # ── B2C Sheet (consumer sales) ──
    ws_b2c = wb.create_sheet("B2C")
    _header_row(ws_b2c, 1, f"GSTR-1 B2C Summary: {date_from} to {date_to}")
    b2c_headers = ["Type", "Place of Supply", "Rate", "Taxable Value", "IGST", "CGST", "SGST", "Total"]
    _styled_headers(ws_b2c, 3, b2c_headers)

    # Aggregate B2C by rate
    b2c_invoices = [inv for inv in invoices if not inv.get("buyer_gstin")]
    rate_totals: dict[int, float] = {}
    for inv in b2c_invoices:
        rate = inv.get("gst_rate", 18)
        taxable = inv.get("taxable_value", inv.get("total_amount", 0) / (1 + rate / 100))
        rate_totals[rate] = rate_totals.get(rate, 0) + taxable

    row = 4
    for rate in sorted(rate_totals.keys()):
        taxable = rate_totals[rate]
        gst = taxable * rate / 100
        ws_b2c.cell(row=row, column=1, value="OE")  # outward exempt / taxable
        ws_b2c.cell(row=row, column=2, value="")
        ws_b2c.cell(row=row, column=3, value=f"{rate}%")
        ws_b2c.cell(row=row, column=4, value=round(taxable, 2))
        ws_b2c.cell(row=row, column=5, value=0)
        ws_b2c.cell(row=row, column=6, value=round(gst / 2, 2))
        ws_b2c.cell(row=row, column=7, value=round(gst / 2, 2))
        ws_b2c.cell(row=row, column=8, value=round(taxable + gst, 2))
        row += 1

    # ── HSN Summary Sheet ──
    ws_hsn = wb.create_sheet("HSN Summary")
    _header_row(ws_hsn, 1, f"HSN-wise Summary: {date_from} to {date_to}")
    hsn_headers = ["HSN Code", "Description", "UQC", "Total Qty", "Taxable Value", "IGST", "CGST", "SGST", "Total Tax"]
    _styled_headers(ws_hsn, 3, hsn_headers)

    # Aggregate by HSN
    hsn_totals: dict[str, dict] = {}
    for inv in invoices:
        for item in inv.get("items", []):
            hsn = item.get("hsn_code", "0000")
            if hsn not in hsn_totals:
                hsn_totals[hsn] = {"description": item.get("description", ""), "qty": 0, "taxable": 0}
            hsn_totals[hsn]["qty"] += item.get("qty", 1)
            hsn_totals[hsn]["taxable"] += item.get("taxable_value", item.get("total", 0))

    row = 4
    for hsn, data in sorted(hsn_totals.items()):
        rate = 18  # default
        gst = data["taxable"] * rate / 100
        ws_hsn.cell(row=row, column=1, value=hsn)
        ws_hsn.cell(row=row, column=2, value=data["description"][:40])
        ws_hsn.cell(row=row, column=3, value="NOS")
        ws_hsn.cell(row=row, column=4, value=data["qty"])
        ws_hsn.cell(row=row, column=5, value=round(data["taxable"], 2))
        ws_hsn.cell(row=row, column=6, value=0)
        ws_hsn.cell(row=row, column=7, value=round(gst / 2, 2))
        ws_hsn.cell(row=row, column=8, value=round(gst / 2, 2))
        ws_hsn.cell(row=row, column=9, value=round(gst, 2))
        row += 1

    _auto_width(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_gstr3b_excel(
    sales_data: dict,
    purchase_data: dict,
    date_from: str,
    date_to: str,
    gstin: str = "",
) -> io.BytesIO:
    """Generate GSTR-3B format Excel (monthly summary return)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B"

    _header_row(ws, 1, f"GSTR-3B Summary Return: {date_from} to {date_to}")
    _header_row(ws, 2, f"GSTIN: {gstin}")

    # 3.1 - Outward supplies
    ws.cell(row=4, column=1, value="3.1 Outward Supplies").font = Font(bold=True, size=12)
    headers_31 = ["Nature of Supplies", "Total Taxable Value", "IGST", "CGST", "SGST", "Cess"]
    _styled_headers(ws, 5, headers_31)

    outward_taxable = sales_data.get("taxable_value", 0)
    outward_gst = sales_data.get("gst_collected", 0)

    ws.cell(row=6, column=1, value="(a) Outward taxable supplies")
    ws.cell(row=6, column=2, value=round(outward_taxable, 2))
    ws.cell(row=6, column=3, value=0)
    ws.cell(row=6, column=4, value=round(outward_gst / 2, 2))
    ws.cell(row=6, column=5, value=round(outward_gst / 2, 2))
    ws.cell(row=6, column=6, value=0)

    # 4 - ITC
    ws.cell(row=9, column=1, value="4. Eligible ITC").font = Font(bold=True, size=12)
    headers_4 = ["Details", "IGST", "CGST", "SGST", "Cess"]
    _styled_headers(ws, 10, headers_4)

    input_gst = purchase_data.get("gst_paid", 0)
    ws.cell(row=11, column=1, value="(A) ITC Available")
    ws.cell(row=11, column=2, value=0)
    ws.cell(row=11, column=3, value=round(input_gst / 2, 2))
    ws.cell(row=11, column=4, value=round(input_gst / 2, 2))

    # 6 - Payment of tax
    ws.cell(row=14, column=1, value="6. Payment of Tax").font = Font(bold=True, size=12)
    net_gst = outward_gst - input_gst
    ws.cell(row=15, column=1, value="Net GST Payable")
    ws.cell(row=15, column=2, value=round(max(0, net_gst), 2))

    _auto_width(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pnl_excel(
    revenue: float,
    cost_of_goods: float,
    gst_collected: float,
    returns_amount: float,
    expenses: dict[str, float] | None = None,
    period: str = "",
    store_name: str = "RetailOS Supermart",
) -> io.BytesIO:
    """Generate detailed P&L statement as Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    expenses = expenses or {}

    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")

    ws.merge_cells("A1:C1")
    ws["A1"] = f"{store_name} — Profit & Loss Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(italic=True)

    # Revenue section
    row = 4
    for col, h in enumerate(["Line Item", "Amount (INR)", "% of Revenue"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    net_revenue = revenue - returns_amount
    gross_profit = net_revenue - cost_of_goods
    total_expenses = sum(expenses.values())
    operating_profit = gross_profit - total_expenses

    rows = [
        ("Gross Revenue", revenue, revenue / revenue * 100 if revenue else 0),
        ("Less: Returns & Refunds", -returns_amount, -returns_amount / revenue * 100 if revenue else 0),
        ("Net Revenue", net_revenue, net_revenue / revenue * 100 if revenue else 0),
        ("", "", ""),
        ("Cost of Goods Sold", -cost_of_goods, -cost_of_goods / revenue * 100 if revenue else 0),
        ("Gross Profit", gross_profit, gross_profit / revenue * 100 if revenue else 0),
        ("", "", ""),
    ]

    for exp_name, exp_amount in expenses.items():
        rows.append((f"  {exp_name}", -exp_amount, -exp_amount / revenue * 100 if revenue else 0))

    if expenses:
        rows.append(("Total Operating Expenses", -total_expenses, -total_expenses / revenue * 100 if revenue else 0))
        rows.append(("", "", ""))

    rows.extend([
        ("Operating Profit", operating_profit, operating_profit / revenue * 100 if revenue else 0),
        ("", "", ""),
        ("GST Collected", gst_collected, ""),
    ])

    row = 5
    for label, amount, pct in rows:
        ws.cell(row=row, column=1, value=label)
        if isinstance(amount, (int, float)) and amount:
            ws.cell(row=row, column=2, value=round(amount, 2))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
        if isinstance(pct, (int, float)) and pct:
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        if label in ("Net Revenue", "Gross Profit", "Operating Profit"):
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    _auto_width(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Helpers ──

def _header_row(ws, row: int, text: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12)


def _styled_headers(ws, row: int, headers: list[str]):
    fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")


def _auto_width(wb):
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)
