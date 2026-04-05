"""Report generators for PDF and Excel exports."""

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

GST_RATES = {
    "Grocery": 0.05, "Dairy": 0.05, "Frozen": 0.12, "Snacks": 0.12,
    "Beverages": 0.12, "Personal Care": 0.18, "Cleaning": 0.18,
    "Baby Care": 0.12, "Bakery": 0.05, "Protein & Health": 0.18,
}


def generate_sales_excel(orders: list[dict], date_from: str, date_to: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header styling
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Sales Report: {date_from} to {date_to}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers = ["Order ID", "Date", "Customer", "Items", "Subtotal", "GST", "Total", "Payment"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 4
    total_revenue = 0
    total_gst = 0
    for order in orders:
        ts = order.get("timestamp", 0)
        dt = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M") if ts else ""
        items_str = ", ".join(f"{i['product_name']} x{i['qty']}" for i in order.get("items", []))
        subtotal = order.get("total_amount", 0) - order.get("gst_amount", 0)
        gst = order.get("gst_amount", 0)

        ws.cell(row=row, column=1, value=order.get("order_id", ""))
        ws.cell(row=row, column=2, value=dt)
        ws.cell(row=row, column=3, value=order.get("customer_name", "Walk-in"))
        ws.cell(row=row, column=4, value=items_str)
        ws.cell(row=row, column=5, value=round(subtotal, 2))
        ws.cell(row=row, column=6, value=round(gst, 2))
        ws.cell(row=row, column=7, value=round(order.get("total_amount", 0), 2))
        ws.cell(row=row, column=8, value=order.get("payment_method", "Cash"))

        total_revenue += order.get("total_amount", 0)
        total_gst += gst
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="TOTALS").font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_revenue - total_gst, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_gst, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_revenue, 2)).font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pnl_pdf(
    revenue: float,
    cost_of_goods: float,
    gst_collected: float,
    returns_amount: float,
    period: str,
    store_name: str = "RetailOS Supermart",
) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f"{store_name} — Profit & Loss Statement", title_style))
    elements.append(Paragraph(f"Period: {period}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    gross_profit = revenue - cost_of_goods
    net_revenue = revenue - returns_amount
    net_profit = net_revenue - cost_of_goods

    data = [
        ["Line Item", "Amount (INR)"],
        ["Gross Revenue", f"₹{revenue:,.2f}"],
        ["Less: Returns & Refunds", f"(₹{returns_amount:,.2f})"],
        ["Net Revenue", f"₹{net_revenue:,.2f}"],
        ["", ""],
        ["Cost of Goods Sold", f"(₹{cost_of_goods:,.2f})"],
        ["Gross Profit", f"₹{gross_profit:,.2f}"],
        ["Gross Margin", f"{(gross_profit / revenue * 100) if revenue else 0:.1f}%"],
        ["", ""],
        ["GST Collected", f"₹{gst_collected:,.2f}"],
        ["", ""],
        ["Net Profit (before tax)", f"₹{net_profit:,.2f}"],
    ]

    table = Table(data, colWidths=[280, 180])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5233")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def generate_gst_excel(orders: list[dict], date_from: str, date_to: str) -> io.BytesIO:
    """GST returns summary grouped by tax slab."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Summary"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"GST Returns Summary: {date_from} to {date_to}"
    ws["A1"].font = Font(bold=True, size=14)

    # Accumulate by category
    category_totals: dict[str, dict] = {}
    for order in orders:
        for item in order.get("items", []):
            cat = item.get("category", "Grocery")
            if cat not in category_totals:
                category_totals[cat] = {"taxable_value": 0, "items_count": 0}
            category_totals[cat]["taxable_value"] += item.get("total", item.get("unit_price", 0) * item.get("qty", 1))
            category_totals[cat]["items_count"] += item.get("qty", 1)

    headers = ["Category", "GST Rate", "Taxable Value", "CGST", "SGST", "Total GST"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row = 4
    grand_taxable = 0
    grand_gst = 0
    for cat, data in sorted(category_totals.items()):
        rate = GST_RATES.get(cat, 0.12)
        taxable = data["taxable_value"]
        gst = taxable * rate
        cgst = gst / 2
        sgst = gst / 2

        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=f"{rate * 100:.0f}%")
        ws.cell(row=row, column=3, value=round(taxable, 2))
        ws.cell(row=row, column=4, value=round(cgst, 2))
        ws.cell(row=row, column=5, value=round(sgst, 2))
        ws.cell(row=row, column=6, value=round(gst, 2))

        grand_taxable += taxable
        grand_gst += gst
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(grand_taxable, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(grand_gst, 2)).font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_inventory_excel(products: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Inventory Report — {date.today().isoformat()}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["SKU", "Product", "Category", "Stock", "Threshold", "Daily Rate", "Days Left", "Status"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx, p in enumerate(products, 4):
        daily = p.get("daily_sales_rate", 0)
        stock = p.get("current_stock", 0)
        days_left = stock / daily if daily > 0 else 999
        status = "Critical" if days_left < 2 else "Warning" if days_left < 5 else "OK"

        ws.cell(row=row_idx, column=1, value=p.get("sku", ""))
        ws.cell(row=row_idx, column=2, value=p.get("product_name", ""))
        ws.cell(row=row_idx, column=3, value=p.get("category", ""))
        ws.cell(row=row_idx, column=4, value=stock)
        ws.cell(row=row_idx, column=5, value=p.get("reorder_threshold", 0))
        ws.cell(row=row_idx, column=6, value=daily)
        ws.cell(row=row_idx, column=7, value=round(days_left, 1))
        cell = ws.cell(row=row_idx, column=8, value=status)
        if status == "Critical":
            cell.font = Font(color="CC0000", bold=True)
        elif status == "Warning":
            cell.font = Font(color="CC8800")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
